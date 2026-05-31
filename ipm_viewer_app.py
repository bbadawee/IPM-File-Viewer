import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import json
import subprocess
import os
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# Mappings configuration
CURRENCY_CODES = {
    "368": { "code": "IQD", "symbol": "د.ع", "decimals": 3, "name": "Iraqi Dinar" },
    "840": { "code": "USD", "symbol": "$", "decimals": 2, "name": "US Dollar" },
    "978": { "code": "EUR", "symbol": "€", "decimals": 2, "name": "Euro" },
    "826": { "code": "GBP", "symbol": "£", "decimals": 2, "name": "British Pound" },
    "784": { "code": "AED", "symbol": "د.إ", "decimals": 2, "name": "UAE Dirham" },
    "986": { "code": "BRL", "symbol": "R$", "decimals": 2, "name": "Brazilian Real" },
    "643": { "code": "RUB", "symbol": "₽", "decimals": 2, "name": "Russian Ruble" },
    "392": { "code": "JPY", "symbol": "¥", "decimals": 0, "name": "Japanese Yen" },
    "156": { "code": "CNY", "symbol": "¥", "decimals": 2, "name": "Chinese Yuan" },
    "356": { "code": "INR", "symbol": "₹", "decimals": 2, "name": "Indian Rupee" },
    "400": { "code": "JOD", "symbol": "د.ا", "decimals": 3, "name": "Jordanian Dinar" },
    "414": { "code": "KWD", "symbol": "د.ك", "decimals": 3, "name": "Kuwaiti Dinar" }
}

MCC_CODES = {
    "5411": { "name": "Grocery Stores & Supermarkets", "category": "Groceries" },
    "5541": { "name": "Service Stations (Gasoline)", "category": "Transportation" },
    "5542": { "name": "Automated Fuel Dispensers", "category": "Transportation" },
    "5812": { "name": "Eating Places, Restaurants", "category": "Dining" },
    "5813": { "name": "Drinking Places (Alcoholic Beverages)", "category": "Dining" },
    "5814": { "name": "Fast Food Restaurants", "category": "Dining" },
    "5912": { "name": "Drug Stores & Pharmacies", "category": "Medical" },
    "5983": { "name": "Fuel Dealers (Coal, Wood, Oil)", "category": "Transportation" },
    "5999": { "name": "Miscellaneous Specialty Retail", "category": "Shopping" },
    "6011": { "name": "Automated Cash Disbursers (ATM)", "category": "Finance" },
    "6012": { "name": "Financial Institutions (Merchandise/Services)", "category": "Finance" },
    "5311": { "name": "Department Stores", "category": "Shopping" },
    "4111": { "name": "Local Passenger Transportation", "category": "Transportation" },
    "4121": { "name": "Taxicabs & Limousines", "category": "Transportation" },
    "4722": { "name": "Travel Agencies & Tour Operators", "category": "Travel" },
    "7011": { "name": "Hotels, Motels & Resorts", "category": "Travel" },
    "7997": { "name": "Membership Clubs (Sports, Recreation)", "category": "Entertainment" },
    "7999": { "name": "Recreation Services", "category": "Entertainment" }
}

EMV_TAGS = {
    "50": { "name": "Application Label", "format": "ascii" },
    "57": { "name": "Track 2 Equivalent Data", "format": "hex" },
    "5A": { "name": "Application PAN", "format": "hex" },
    "5F20": { "name": "Cardholder Name", "format": "ascii" },
    "5F24": { "name": "Application Expiration Date", "format": "bcd_date" },
    "5F25": { "name": "Application Effective Date", "format": "bcd_date" },
    "5F28": { "name": "Issuer Country Code", "format": "numeric" },
    "5F2D": { "name": "Language Preference", "format": "ascii" },
    "5F34": { "name": "Application PAN Sequence Number", "format": "hex" },
    "5F2A": { "name": "Transaction Currency Code", "format": "numeric" },
    "82": { "name": "Application Interchange Profile (AIP)", "format": "hex" },
    "84": { "name": "Dedicated File Name (DF Name / AID)", "format": "hex" },
    "8A": { "name": "Authorization Response Code", "format": "ascii" },
    "95": { "name": "Terminal Verification Results (TVR)", "format": "hex" },
    "9A": { "name": "Transaction Date", "format": "bcd_date" },
    "9C": { "name": "Transaction Type", "format": "hex" },
    "9F02": { "name": "Amount, Authorized (Numeric)", "format": "numeric_amount" },
    "9F03": { "name": "Amount, Other (Numeric)", "format": "numeric_amount" },
    "9F10": { "name": "Issuer Application Data (IAD)", "format": "hex" },
    "9F1A": { "name": "Terminal Country Code", "format": "numeric" },
    "9F26": { "name": "Application Cryptogram (AC)", "format": "hex" },
    "9F27": { "name": "Cryptogram Information Data (CID)", "format": "hex" },
    "9F33": { "name": "Terminal Capabilities", "format": "hex" },
    "9F34": { "name": "Cardholder Verification Method (CVM) Results", "format": "hex" },
    "9F36": { "name": "Application Transaction Counter (ATC)", "format": "hex" },
    "9F37": { "name": "Unpredictable Number", "format": "hex" },
    "9F42": { "name": "Application Currency Code", "format": "numeric" },
    "9F6E": { "name": "Third Party Data", "format": "hex" },
    "9B": { "name": "Transaction Status Information (TSI)", "format": "hex" },
    "8E": { "name": "Cardholder Verification Method (CVM) List", "format": "hex" },
    "9F12": { "name": "Application Preferred Name", "format": "ascii" }
}

class IPMViewerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("IPMView")
        self.geometry("1000x700")
        
        # Native theme config
        self.style = ttk.Style(self)
        if "vista" in self.style.theme_names():
            self.style.theme_use("vista")
        elif "xpnative" in self.style.theme_names():
            self.style.theme_use("xpnative")
            
        self.style.configure("Treeview", rowheight=24) # Spaced layout row height
            
        # App state variables
        self.all_transactions = []
        self.filtered_transactions = []
        self.is_pan_masked = True
        self.filter_visible = True
        self.loaded_file_path = ""
        
        self.setup_menu()
        self.setup_layout()
        
    def setup_menu(self):
        menu_bar = tk.Menu(self)
        
        # File Menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Open IPM File...", command=self.open_file_dialog, accelerator="Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Export Filtered List to Excel...", command=self.export_to_excel)
        file_menu.add_command(label="Generate PDF Executive Summary...", command=self.generate_pdf_report)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)
        menu_bar.add_cascade(label="File", menu=file_menu)
        
        # Config Menu
        config_menu = tk.Menu(menu_bar, tearoff=0)
        config_menu.add_checkbutton(label="Mask Card PANs", command=self.toggle_pan_masking, accelerator="Ctrl+M")
        # Initialize checkbutton state matching default setting
        config_menu.entryconfigure("Mask Card PANs", variable=tk.BooleanVar(value=True))
        menu_bar.add_cascade(label="Config", menu=config_menu)
        
        # Help Menu
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about)
        menu_bar.add_cascade(label="Help", menu=help_menu)
        
        self.config(menu=menu_bar)
        
        # Bind keyboard shortcuts
        self.bind("<Control-o>", lambda event: self.open_file_dialog())
        self.bind("<Control-m>", lambda event: self.toggle_pan_masking())
        self.bind("<Control-f>", lambda event: self.search_entry.focus_set())
        self.bind("<Escape>", lambda event: self.reset_filters())

    def setup_layout(self):
        # 1. Top Collapsible Filter Frame
        self.filter_header_frame = ttk.Frame(self)
        self.filter_header_frame.pack(fill="x", padx=5, pady=2)
        
        self.filter_toggle_btn = ttk.Button(self.filter_header_frame, text="▼ Filter", command=self.toggle_filter_panel, width=10)
        self.filter_toggle_btn.pack(side="left")
        
        self.filter_controls_frame = ttk.LabelFrame(self, text="")
        self.filter_controls_frame.pack(fill="x", padx=10, pady=5)
        
        # Grid layout inside filter controls
        ttk.Label(self.filter_controls_frame, text="Search (PAN, Merchant, RRN...):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.search_entry = ttk.Entry(self.filter_controls_frame, width=25)
        self.search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.search_entry.bind("<KeyRelease>", lambda e: self.apply_filters())
        
        ttk.Label(self.filter_controls_frame, text="Brand:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.brand_combo = ttk.Combobox(self.filter_controls_frame, values=["All", "Mastercard", "Visa", "Other"], width=12, state="readonly")
        self.brand_combo.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.brand_combo.set("All")
        self.brand_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_filters())
        
        ttk.Label(self.filter_controls_frame, text="MTI:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.mti_combo = ttk.Combobox(self.filter_controls_frame, values=["All"], width=10, state="readonly")
        self.mti_combo.grid(row=0, column=5, padx=5, pady=5, sticky="w")
        self.mti_combo.set("All")
        self.mti_combo.bind("<<ComboboxSelected>>", lambda e: self.apply_filters())
        
        self.reset_btn = ttk.Button(self.filter_controls_frame, text="Reset", command=self.reset_filters, width=8)
        self.reset_btn.grid(row=0, column=6, padx=10, pady=5, sticky="w")
        
        self.expand_btn = ttk.Button(self.filter_controls_frame, text="Expand All", command=self.expand_all, width=12)
        self.expand_btn.grid(row=0, column=7, padx=5, pady=5, sticky="w")
        
        self.collapse_btn = ttk.Button(self.filter_controls_frame, text="Collapse All", command=self.collapse_all, width=12)
        self.collapse_btn.grid(row=0, column=8, padx=5, pady=5, sticky="w")

        # 1b. File Overview / Dashboard Summary Ribbon (just below filters)
        self.summary_frame = ttk.LabelFrame(self, text="File Overview (Filtered)")
        self.summary_frame.pack(fill="x", padx=10, pady=5)
        
        self.lbl_summary_tx = ttk.Label(self.summary_frame, text="Total Transactions: -", font=("Segoe UI", 9, "bold"))
        self.lbl_summary_tx.pack(side="left", padx=15, pady=5)
        
        self.lbl_summary_vol = ttk.Label(self.summary_frame, text="Volume: -", font=("Segoe UI", 9, "bold"))
        self.lbl_summary_vol.pack(side="left", padx=15, pady=5)
        
        self.lbl_summary_mix = ttk.Label(self.summary_frame, text="Scheme Mix: -", font=("Segoe UI", 9, "bold"))
        self.lbl_summary_mix.pack(side="left", padx=15, pady=5)
        
        # 2. Main Tree Table Grid
        self.tree_container = ttk.Frame(self)
        self.tree_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.tree = ttk.Treeview(self.tree_container, columns=("Value", "Meaning"), selectmode="browse")
        
        # Bind double-click row to expand/collapse
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # Tree Headers
        self.tree.heading("#0", text="Field", anchor="w")
        self.tree.heading("Value", text="Value", anchor="w")
        self.tree.heading("Meaning", text="Meaning", anchor="w")
        
        # Column formatting
        self.tree.column("#0", width=380, minwidth=250, stretch=True)
        self.tree.column("Value", width=250, minwidth=150, stretch=True)
        self.tree.column("Meaning", width=350, minwidth=200, stretch=True)
        
        # Scrollbars
        self.tree_scroll_y = ttk.Scrollbar(self.tree_container, orient="vertical", command=self.tree.yview)
        self.tree_scroll_x = ttk.Scrollbar(self.tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        self.tree_scroll_y.pack(side="right", fill="y")
        self.tree_scroll_x.pack(side="bottom", fill="x")
        
        # Configure alternating row colors
        self.tree.tag_configure("even", background="#f3f4f6")
        self.tree.tag_configure("odd", background="#ffffff")
        
        # 3. Bottom Status Bar
        self.status_bar = ttk.Label(self, text="Ready. Open an IPM clearing file (File -> Open or Ctrl+O) to begin.", relief="sunken", anchor="w")
        self.status_bar.pack(fill="x", side="bottom")

    def toggle_filter_panel(self):
        if self.filter_visible:
            self.filter_controls_frame.pack_forget()
            self.filter_toggle_btn.config(text="▸ Filter")
            self.filter_visible = False
        else:
            self.filter_controls_frame.pack(fill="x", padx=10, pady=5, after=self.filter_header_frame)
            self.filter_toggle_btn.config(text="▼ Filter")
            self.filter_visible = True

    def open_file_dialog(self):
        file_path = filedialog.askopenfilename(
            title="Open IPM Clearing File",
            filetypes=[("IPM / Clearing Files", "*.ipm;*.dat"), ("All Files", "*.*")]
        )
        if file_path:
            self.load_ipm_file(file_path)

    def load_ipm_file(self, file_path):
        # Update status
        self.status_bar.config(text=f"Loading and converting: {os.path.basename(file_path)}...")
        self.update()
        
        exe_path = resource_path("ipm2json.exe")
        if not os.path.exists(exe_path):
            messagebox.showerror("Error", f"Conversion tool 'ipm2json.exe' not found at:\n{exe_path}")
            self.status_bar.config(text="Error: Conversion tool not found.")
            return

        try:
            # Run ipm2json.exe with -u UTF-8 output
            result = subprocess.run([exe_path, "-u", file_path], capture_output=True, text=True, check=True)
            
            # Load JSON content
            transactions_data = json.loads(result.stdout)
            
            # Update state
            self.all_transactions = transactions_data
            self.loaded_file_path = file_path
            
            # Update Window title to show file name matching screenshot
            self.title(f"IPMView - {os.path.basename(file_path)}")
            
            # Dynamically populate MTI filter options
            mti_set = set()
            for tx in self.all_transactions:
                if "mti" in tx:
                    mti_set.add(tx["mti"])
            self.mti_combo["values"] = ["All"] + sorted(list(mti_set))
            self.mti_combo.set("All")
            
            self.reset_filters()
            
        except subprocess.CalledProcessError as err:
            err_msg = err.stderr or err.output or str(err)
            messagebox.showerror("Conversion Error", f"Failed to convert IPM file:\n{err_msg}")
            self.status_bar.config(text="Error converting IPM file.")
        except json.JSONDecodeError as err:
            messagebox.showerror("Parse Error", f"Failed to parse conversion output as JSON:\n{str(err)}")
            self.status_bar.config(text="Error parsing conversion output.")
        except Exception as err:
            messagebox.showerror("Error", f"An unexpected error occurred:\n{str(err)}")
            self.status_bar.config(text="An unexpected error occurred.")

    def reset_filters(self):
        self.search_entry.delete(0, tk.END)
        self.brand_combo.set("All")
        self.mti_combo.set("All")
        self.apply_filters()

    def toggle_pan_masking(self):
        self.is_pan_masked = not self.is_pan_masked
        # Reflect in tree re-rendering
        self.populate_tree()

    def show_about(self):
        messagebox.showinfo("About IPMView", 
                            "IPM Clearing File Viewer\n\n"
                            "A Windows desktop utility for Mastercard and ISO-8583 clearing presentment and fee collection files.\n"
                            "Uses native Tkinter tree-grid layout and maps MCCs, currencies, and EMV tags.")

    def export_to_excel(self):
        if not self.filtered_transactions:
            messagebox.showwarning("Export Warning", "No transaction data available to export.")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not file_path:
            return
            
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions Log"
            
            # Define headers
            headers = [
                "Index", "MTI", "Function Code", "PAN (DE2)", "Card Brand", 
                "Raw Amount (DE4)", "Formatted Amount", "Currency (DE49)", 
                "Local Time (DE12)", "Expiration (DE14)", "MCC (DE26)", 
                "MCC Category", "Merchant Name", "Merchant City", "Merchant Country", 
                "Approval Code (DE38)", "Retrieval Ref Num (DE37)", 
                "Acquirer Ref Num (DE31)", "Terminal ID (DE41)", "Merchant ID (DE42)"
            ]
            ws.append(headers)
            
            # Style header row
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            # Populate rows
            for idx, tx in enumerate(self.filtered_transactions):
                mti = tx.get("mti", "")
                de24 = tx.get("de24", "")
                
                # PAN
                raw_pan = tx.get("de2", "")
                pan = self.mask_pan(raw_pan) if self.is_pan_masked else self.format_pan_spacing(raw_pan)
                brand = self.get_card_brand(raw_pan)
                
                # Amount
                raw_amt = tx.get("de4", "")
                curr_code = tx.get("de49", "")
                formatted_amt = self.format_amount(raw_amt, curr_code)
                
                # Date
                de12 = tx.get("de12", {})
                s1_date = de12.get("s1", "")
                s2_time = de12.get("s2", "")
                formatted_time = ""
                if len(s1_date) == 6 and len(s2_time) == 6:
                    formatted_time = f"20{s1_date[0:2]}-{s1_date[2:4]}-{s1_date[4:6]} {s2_time[0:2]}:{s2_time[2:4]}:{s2_time[4:6]}"
                
                de14 = tx.get("de14", "")
                formatted_exp = ""
                if len(de14) == 4:
                    formatted_exp = f"{de14[2:4]}/20{de14[0:2]}"
                    
                # MCC
                mcc = tx.get("de26", "")
                mcc_meta = MCC_CODES.get(mcc, {"name": "", "category": ""})
                mcc_cat = mcc_meta.get("category", "")
                
                # Merchant
                name, city, country = self.get_merchant_name_and_city(tx)
                
                # Reference codes
                de38 = tx.get("de38", "")
                de37 = tx.get("de37", "")
                
                # ARN
                de31 = tx.get("de31", {})
                arn = ""
                if isinstance(de31, dict):
                    arn = "".join([str(de31.get(k, "")) for k in sorted(de31.keys())])
                    
                de41 = tx.get("de41", "")
                de42 = tx.get("de42", "").strip() if tx.get("de42") else ""
                
                row_data = [
                    idx + 1, mti, de24, pan, brand,
                    raw_amt, formatted_amt, curr_code,
                    formatted_time, formatted_exp, mcc,
                    mcc_cat, name, city, country,
                    de38, de37, arn, de41, de42
                ]
                ws.append(row_data)
                
                # Style row borders and alignments
                row_num = idx + 2
                for col_num in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = Font(name="Segoe UI", size=10)
                    cell.border = thin_border
                    
                    # Alignments
                    if col_num in [1, 2, 3, 5, 8, 9, 10, 11, 16, 17, 18, 19]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_num in [6]:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                        
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Excel sheet exported successfully to:\n{file_path}")
            self.status_bar.config(text=f"Exported list to Excel: {os.path.basename(file_path)}")
            
        except ImportError:
            messagebox.showerror("Dependency Error", "Please install 'openpyxl' to export Excel files:\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save Excel file:\n{str(e)}")

    def generate_pdf_report(self):
        if not self.filtered_transactions:
            messagebox.showwarning("Export Warning", "No transaction data available to generate report.")
            return
            
        file_path = filedialog.asksaveasfilename(
            title="Save PDF Report",
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not file_path:
            return
            
        try:
            from fpdf import FPDF
            from datetime import datetime
            
            filename_base = os.path.basename(self.loaded_file_path)
            
            # Define PDF class inline to keep it self-contained
            class ExecutiveSummaryPDF(FPDF):
                def __init__(self, filename_source):
                    super().__init__()
                    self.filename_source = filename_source
                    
                def header(self):
                    self.set_font("Helvetica", "B", 8)
                    self.set_text_color(100, 100, 100)
                    self.cell(0, 10, f"IPM Clearing Report - Source: {self.filename_source}", 0, 0, "L")
                    self.set_font("Helvetica", "", 8)
                    self.cell(0, 10, "CONFIDENTIAL / INTERNAL USE", 0, 0, "R")
                    self.ln(8)
                    self.set_draw_color(220, 220, 220)
                    self.line(10, 18, 200, 18)
                    self.ln(5)
                    
                def footer(self):
                    self.set_y(-15)
                    self.set_font("Helvetica", "I", 8)
                    self.set_text_color(150, 150, 150)
                    self.set_draw_color(220, 220, 220)
                    self.line(10, 282, 200, 282)
                    self.cell(0, 10, f"IPMView Desktop Audit Tool", 0, 0, "L")
                    self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "R")
                    
            pdf = ExecutiveSummaryPDF(filename_base)
            pdf.set_auto_page_break(auto=True, margin=20)
            pdf.add_page()
            
            # --- COVER PAGE HEADER ---
            pdf.set_font("Helvetica", "B", 18)
            pdf.set_text_color(31, 73, 125) # Deep Blue
            pdf.cell(0, 12, "IPM Clearing File Audit Report", 0, 1, "L")
            
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(128, 128, 128)
            pdf.cell(0, 5, f"Date Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, "L")
            pdf.cell(0, 5, f"Clearing Source File: {filename_base}", 0, 1, "L")
            pdf.ln(10)
            
            # --- EXECUTIVE SUMMARY SUMMARY GRID ---
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(31, 73, 125)
            pdf.cell(0, 8, "1. Executive Summary", 0, 1, "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            
            total_tx = len(self.filtered_transactions)
            total_file_tx = len(self.all_transactions)
            pdf.multi_cell(0, 5, f"This audit report provides an executive summary of transactions contained inside the Mastercard clearing file. Out of a total {total_file_tx} records, {total_tx} transactions match the active search filters and are detailed below.")
            pdf.ln(5)
            
            # Volume & counts summary tables
            volumes = {}
            mti_counts = {}
            scheme_counts = {"Mastercard": 0, "Visa": 0, "Other": 0}
            
            for tx in self.filtered_transactions:
                curr = tx.get("de49", "Unknown")
                raw_amt = float(tx.get("de4", 0)) if tx.get("de4") else 0
                meta = CURRENCY_CODES.get(curr, {"decimals": 2, "code": curr})
                real_amt = raw_amt / (10 ** meta.get("decimals", 2))
                volumes[curr] = volumes.get(curr, 0) + real_amt
                
                mti = tx.get("mti", "Unknown")
                mti_counts[mti] = mti_counts.get(mti, 0) + 1
                
                brand = self.get_card_brand(tx.get("de2", ""))
                scheme_counts[brand] = scheme_counts.get(brand, 0) + 1
                
            # Table 1: Volume Summary
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(255, 255, 255)
            pdf.set_fill_color(31, 73, 125)
            pdf.cell(60, 7, "Currency Code", 1, 0, "C", True)
            pdf.cell(80, 7, "Clearing Volume", 1, 0, "C", True)
            pdf.cell(50, 7, "Transaction Count", 1, 1, "C", True)
            
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(50, 50, 50)
            
            for idx, curr in enumerate(sorted(volumes.keys())):
                meta = CURRENCY_CODES.get(curr, {"code": curr, "decimals": 2, "name": "Unknown"})
                vol = volumes[curr]
                formatted_vol = f"{vol:,.{meta.get('decimals', 2)}f} {meta.get('code')}"
                
                # Count
                curr_tx_count = sum(1 for tx in self.filtered_transactions if tx.get("de49") == curr)
                
                fill = idx % 2 == 1
                pdf.set_fill_color(240, 244, 248)
                pdf.cell(60, 6, f"{meta.get('name')} ({curr})", 1, 0, "L", fill)
                pdf.cell(80, 6, formatted_vol, 1, 0, "R", fill)
                pdf.cell(50, 6, str(curr_tx_count), 1, 1, "C", fill)
            pdf.ln(8)
            
            # Table 2: MTI and Card Schemes
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(255, 255, 255)
            pdf.set_fill_color(31, 73, 125)
            pdf.cell(95, 7, "Message Type (MTI)", 1, 0, "C", True)
            pdf.cell(95, 7, "Card Brand Scheme Mix", 1, 1, "C", True)
            
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(50, 50, 50)
            
            # Combine keys to draw parallel rows
            mti_keys = sorted(mti_counts.keys())
            scheme_keys = sorted(scheme_counts.keys())
            max_rows = max(len(mti_keys), len(scheme_keys))
            
            for i in range(max_rows):
                fill = i % 2 == 1
                pdf.set_fill_color(240, 244, 248)
                
                # MTI Col
                if i < len(mti_keys):
                    mti = mti_keys[i]
                    count = mti_counts[mti]
                    mti_desc = self.get_mti_short_name(mti)
                    mti_text = f"{mti} - {mti_desc}: {count} ({count/total_tx*100:.1f}%)"
                else:
                    mti_text = ""
                pdf.cell(95, 6, mti_text, 1, 0, "L", fill)
                
                # Scheme Col
                if i < len(scheme_keys):
                    brand = scheme_keys[i]
                    count = scheme_counts[brand]
                    scheme_text = f"{brand}: {count} ({count/total_tx*100:.1f}%)" if count > 0 else ""
                else:
                    scheme_text = ""
                pdf.cell(95, 6, scheme_text, 1, 1, "L", fill)
                
            pdf.ln(8)
            
            # --- TOP MERCHANT LEADERBOARD ---
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(31, 73, 125)
            pdf.cell(0, 8, "2. Top Merchants", 0, 1, "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 5, "Ranked by total volume processed in the clearing file (Top 5):", 0, 1, "L")
            pdf.ln(3)
            
            # Calculate top merchants
            merchant_stats = {}
            for tx in self.filtered_transactions:
                name, city, country = self.get_merchant_name_and_city(tx)
                curr = tx.get("de49", "Unknown")
                raw_amt = float(tx.get("de4", 0)) if tx.get("de4") else 0
                meta = CURRENCY_CODES.get(curr, {"decimals": 2, "code": curr})
                real_amt = raw_amt / (10 ** meta.get("decimals", 2))
                
                key = (name, city, country, curr)
                if key not in merchant_stats:
                    merchant_stats[key] = {"count": 0, "volume": 0.0}
                merchant_stats[key]["count"] += 1
                merchant_stats[key]["volume"] += real_amt
                
            sorted_merchants = sorted(merchant_stats.items(), key=lambda item: item[1]["volume"], reverse=True)[:5]
            
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(255, 255, 255)
            pdf.set_fill_color(31, 73, 125)
            pdf.cell(10, 7, "No", 1, 0, "C", True)
            pdf.cell(85, 7, "Merchant Name & Location", 1, 0, "C", True)
            pdf.cell(20, 7, "Count", 1, 0, "C", True)
            pdf.cell(75, 7, "Total Volume", 1, 1, "C", True)
            
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(50, 50, 50)
            for idx, (key, stats) in enumerate(sorted_merchants):
                name, city, country, curr = key
                loc = f"{city + ', ' if city else ''}{country}"
                display_name = f"{name} ({loc})"
                
                meta = CURRENCY_CODES.get(curr, {"code": curr, "decimals": 2})
                vol_str = f"{stats['volume']:,.{meta['decimals']}f} {meta['code']}"
                
                fill = idx % 2 == 1
                pdf.set_fill_color(240, 244, 248)
                pdf.cell(10, 6, str(idx + 1), 1, 0, "C", fill)
                pdf.cell(85, 6, display_name[:42], 1, 0, "L", fill)
                pdf.cell(20, 6, str(stats['count']), 1, 0, "C", fill)
                pdf.cell(75, 6, vol_str, 1, 1, "R", fill)
                
            pdf.ln(8)
            
            # --- APPENDIX LOG TABLE ---
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(31, 73, 125)
            pdf.cell(0, 8, "3. Appendix: Transaction Audit Log (First 50)", 0, 1, "L")
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 5, "Detailed listing of transaction occurrences in sequence order:", 0, 1, "L")
            pdf.ln(3)
            
            # Table Headers
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(255, 255, 255)
            pdf.set_fill_color(31, 73, 125)
            pdf.cell(32, 6, "Card Number (PAN)", 1, 0, "C", True)
            pdf.cell(18, 6, "MTI/Func", 1, 0, "C", True)
            pdf.cell(30, 6, "Local Timestamp", 1, 0, "C", True)
            pdf.cell(50, 6, "Merchant Name & City", 1, 0, "C", True)
            pdf.cell(15, 6, "Auth", 1, 0, "C", True)
            pdf.cell(45, 6, "Clearing Amount", 1, 1, "C", True)
            
            pdf.set_font("Helvetica", "", 8)
            pdf.set_text_color(50, 50, 50)
            
            # Print first 50 transactions
            for idx, tx in enumerate(self.filtered_transactions[:50]):
                raw_pan = tx.get("de2", "")
                pan = self.mask_pan(raw_pan) if self.is_pan_masked else self.format_pan_spacing(raw_pan)
                
                mti = tx.get("mti", "")
                de24 = tx.get("de24", "")
                mti_func = f"{mti}/{de24}"
                
                # Date
                de12 = tx.get("de12", {})
                s1_date = de12.get("s1", "")
                s2_time = de12.get("s2", "")
                time_str = f"20{s1_date[0:2]}-{s1_date[2:4]}-{s1_date[4:6]}" if s1_date else ""
                
                name, city, country = self.get_merchant_name_and_city(tx)
                merch_str = f"{name} ({city})" if city else name
                
                auth = tx.get("de38", "")
                
                # Amount
                raw_amt = tx.get("de4", "")
                curr_code = tx.get("de49", "")
                formatted_amt = self.format_amount(raw_amt, curr_code)
                
                fill = idx % 2 == 1
                pdf.set_fill_color(240, 244, 248)
                
                pdf.cell(32, 5.5, pan, 1, 0, "C", fill)
                pdf.cell(18, 5.5, mti_func, 1, 0, "C", fill)
                pdf.cell(30, 5.5, time_str, 1, 0, "C", fill)
                pdf.cell(50, 5.5, merch_str[:28], 1, 0, "L", fill)
                pdf.cell(15, 5.5, auth, 1, 0, "C", fill)
                pdf.cell(45, 5.5, formatted_amt, 1, 1, "R", fill)
                
            pdf.output(file_path)
            messagebox.showinfo("Report Successful", f"PDF summary report saved successfully to:\n{file_path}")
            self.status_bar.config(text=f"Generated PDF Report: {os.path.basename(file_path)}")
            
        except ImportError:
            messagebox.showerror("Dependency Error", "Please install 'fpdf2' to export PDF files:\npip install fpdf2")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save PDF Report:\n{str(e)}")

    def on_tree_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.item(item, open=not self.tree.item(item, "open"))

    def expand_all(self):
        for item in self.tree.get_children():
            self.tree.item(item, open=True)
            # Expand sub-nodes
            for sub_item in self.tree.get_children(item):
                self.tree.item(sub_item, open=True)

    def collapse_all(self):
        for item in self.tree.get_children():
            self.tree.item(item, open=False)

    def update_summary_stats(self):
        txs = self.filtered_transactions
        total = len(txs)
        self.lbl_summary_tx.config(text=f"Total Transactions: {total}")
        
        # Calculate stats
        volumes = {}
        mc_count = 0
        visa_count = 0
        other_count = 0
        
        for tx in txs:
            curr = tx.get("de49", "Unknown")
            raw_amt = float(tx.get("de4", 0)) if tx.get("de4") else 0
            meta = CURRENCY_CODES.get(curr, {"decimals": 2, "code": curr})
            decimals = meta.get("decimals", 2)
            real_amt = raw_amt / (10 ** decimals)
            
            volumes[curr] = volumes.get(curr, 0) + real_amt
            
            brand = self.get_card_brand(tx.get("de2", ""))
            if brand == "Mastercard":
                mc_count += 1
            elif brand == "Visa":
                visa_count += 1
            else:
                other_count += 1
                
        vol_parts = []
        for curr, vol in volumes.items():
            meta = CURRENCY_CODES.get(curr, {"code": curr, "decimals": 2})
            code = meta.get("code", curr)
            vol_parts.append(f"{vol:,.{meta.get('decimals', 2)}f} {code}")
            
        vol_text = " | ".join(vol_parts) if vol_parts else "0.00"
        self.lbl_summary_vol.config(text=f"Volume: {vol_text}")
        
        mix_text = f"Mastercard: {mc_count} | Visa: {visa_count}"
        if other_count > 0:
            mix_text += f" | Other: {other_count}"
        self.lbl_summary_mix.config(text=f"Scheme Mix: {mix_text}")

    def apply_filters(self):
        search_query = self.search_entry.get().strip().lower()
        selected_brand = self.brand_combo.get()
        selected_mti = self.mti_combo.get()
        
        self.filtered_transactions = []
        
        for tx in self.all_transactions:
            # 1. MTI filter
            if selected_mti != "All" and tx.get("mti") != selected_mti:
                continue
                
            # 2. Brand filter
            brand = self.get_card_brand(tx.get("de2", ""))
            if selected_brand != "All":
                if selected_brand == "Other" and (brand == "Visa" or brand == "Mastercard"):
                    continue
                elif selected_brand != "Other" and brand != selected_brand:
                    continue
            
            # 3. Search query filter (matches PAN, merchant name in DE43, RRN, ARN, Approval Code)
            if search_query:
                pan = tx.get("de2", "")
                rrn = tx.get("de37", "")
                arn = tx.get("de31", {}).get("s4", "") if isinstance(tx.get("de31"), dict) else ""
                auth_code = tx.get("de38", "")
                
                # Merchant name
                name, city, country = self.get_merchant_name_and_city(tx)
                name_l = name.lower()
                city_l = city.lower()
                
                matches = (search_query in pan or 
                           search_query in rrn or 
                           search_query in arn or 
                           search_query in auth_code or 
                           search_query in name_l or 
                           search_query in city_l)
                if not matches:
                    continue
                    
            self.filtered_transactions.append(tx)
            
        self.update_summary_stats()
        self.populate_tree()

    def populate_tree(self):
        # Clear existing items
        self.tree.delete(*self.tree.get_children())
        
        tx_count = len(self.filtered_transactions)
        total_count = len(self.all_transactions)
        self.status_bar.config(text=f"Loaded {tx_count} of {total_count} transactions.")
        
        if tx_count == 0:
            return
            
        # Sorting key for fields: numerical order for DEs, then PDS, then others
        def sort_fields_key(field):
            if field.startswith("de") and field[2:].isdigit():
                return (0, int(field[2:]))
            elif field.startswith("pds") and field[3:].isdigit():
                return (1, int(field[3:]))
            return (2, field)

        # Insert transactions as parent root nodes
        for idx, tx in enumerate(self.filtered_transactions):
            mti = tx.get("mti", "Unknown")
            de24 = tx.get("de24", "")
            
            # Root Node Header: e.g. 1240/200 - First Presentment
            desc = self.get_mti_short_name(mti)
            if de24:
                fn_desc = self.get_function_code_desc(de24)
                root_text = f"{mti}/{de24} - {fn_desc}"
            else:
                root_text = f"{mti} - {desc}"
                
            # If it's a file header 1644/697
            if mti == "1644" and de24 == "697":
                root_text = f"{mti}/{de24} - File Header"
                
            tag = "even" if idx % 2 == 0 else "odd"
            parent_id = self.tree.insert("", "end", text=root_text, values=("", ""), tags=(tag,))
            
            # Sort transaction keys
            fields = sorted([k for k in tx.keys() if k not in ["mti"]], key=sort_fields_key)
            
            for field_key in fields:
                val = tx[field_key]
                label, fmt_val, meaning = self.get_field_info(tx, field_key, val)
                
                # Check if it's a compound field (dict)
                if isinstance(val, dict):
                    sub_node_id = self.tree.insert(parent_id, "end", text=label, values=(fmt_val, meaning))
                    
                    # Add child items inside compound fields
                    if field_key == "de55": # EMV Tags
                        self.populate_emv_tags(sub_node_id, val)
                    else:
                        # Standard subfields (s1, s2, s3...)
                        self.populate_standard_subfields(sub_node_id, field_key, val)
                else:
                    self.tree.insert(parent_id, "end", text=label, values=(fmt_val, meaning))

    def populate_standard_subfields(self, node_id, parent_key, val_dict):
        # Helper to define labels for subfields to make details rich
        subfield_label_maps = {
            "de3": {
                "s1": "s1 - Transaction Type",
                "s2": "s2 - Source Account",
                "s3": "s3 - Destination Account"
            },
            "de12": {
                "s1": "s1 - Local Date (YYMMDD)",
                "s2": "s2 - Local Time (HHMMSS)"
            },
            "de22": {
                "s1": "s1 - PAN Entry Mode",
                "s2": "s2 - PIN Entry Mode",
                "s3": "s3 - Card Unit Capacity",
                "s4": "s4 - Card Holder Verification Capability",
                "s5": "s5 - Terminal Verification Capability",
                "s6": "s6 - POS Operating Environment",
                "s7": "s7 - Cardholder Present Indicator",
                "s8": "s8 - Card Present Indicator",
                "s9": "s9 - Card Input Mode",
                "s10": "s10 - Cardholder Auth Method",
                "s11": "s11 - Cardholder Auth Entity",
                "s12": "s12 - Card Data Output Capability"
            },
            "de43": {
                "s1": "s1 - Billing Name",
                "s2": "s2 - Name & Location",
                "s3": "s3 - City Name",
                "s4": "s4 - State/Province",
                "s5": "s5 - Postal Code",
                "s6": "s6 - Country Code"
            }
        }

        keys = sorted(val_dict.keys())
        for sub_key in keys:
            sub_val = val_dict[sub_key]
            
            # Custom label mapping
            label = f"{sub_key}"
            if parent_key in subfield_label_maps and sub_key in subfield_label_maps[parent_key]:
                label = subfield_label_maps[parent_key][sub_key]
            else:
                label = f"{sub_key} - Subfield"
                
            meaning = ""
            # Simple decoding on subfield if relevant
            if parent_key == "de3" and sub_key == "s1":
                if sub_val == "00": meaning = "Goods & Services / Purchase"
                elif sub_val == "01": meaning = "Cash Withdrawal"
                elif sub_val == "20": meaning = "Refund / Return"
                
            self.tree.insert(node_id, "end", text=label, values=(str(sub_val), meaning))

    def populate_emv_tags(self, node_id, emv_dict):
        keys = sorted(emv_dict.keys())
        for key in keys:
            tag_id = key.replace("tag", "").upper()
            tag_meta = EMV_TAGS.get(tag_id, {"name": f"EMV Tag {tag_id}", "format": "hex"})
            label = f"Tag {tag_id} - {tag_meta.get('name')}"
            hex_val = emv_dict[key]
            
            # Decode meaning
            tag_format = tag_meta.get("format", "hex")
            meaning = ""
            if tag_format == "ascii":
                meaning = self.hex_to_ascii(hex_val)
            elif tag_format == "bcd_date":
                meaning = self.parse_bcd_date(hex_val)
            elif tag_format == "numeric_amount":
                meaning = self.parse_emv_amount(hex_val)
            elif tag_format == "numeric":
                try:
                    meaning = str(int(hex_val, 10))
                except ValueError:
                    meaning = hex_val
            
            self.tree.insert(node_id, "end", text=label, values=(hex_val, meaning))

    # Helper algorithms for parsing ISO elements
    def get_field_info(self, tx, field_key, value):
        label = field_key.upper()
        fmt_val = ""
        meaning = ""
        
        field_labels = {
            "de2": "DE2 - Primary Account Number (PAN)",
            "de3": "DE3 - Processing Code",
            "de4": "DE4 - Amount, Transaction",
            "de12": "DE12 - Date And Time, Local Transaction",
            "de14": "DE14 - Date, Expiration (YYMM)",
            "de22": "DE22 - Point Of Service Data Code",
            "de24": "DE24 - Function Code",
            "de26": "DE26 - Card Acceptor Business Code (MCC)",
            "de31": "DE31 - Acquirer Reference Data",
            "de32": "DE32 - Acquiring Institution Id Code",
            "de33": "DE33 - Forwarding Institution Id Code",
            "de37": "DE37 - Retrieval Reference Number",
            "de38": "DE38 - Approval Code",
            "de41": "DE41 - Card Acceptor Terminal Id",
            "de42": "DE42 - Card Acceptor Id Code",
            "de43": "DE43 - Card Acceptor Name/Location",
            "de49": "DE49 - Transaction Currency Code",
            "de55": "DE55 - Integrated Circuit Card (ICC) Data",
            "de71": "DE71 - Message Number",
            "de94": "DE94 - Transaction Origination Institution Id Code",
        }
        
        if field_key in field_labels:
            label = field_labels[field_key]
        elif field_key.startswith("pds"):
            pds_num = field_key[3:]
            label = f"PDS {pds_num} - Private Data Subelement"
            
        # Parse compound values vs leaf values
        if isinstance(value, dict):
            # Compound field
            sub_keys = sorted(value.keys())
            if all(k.startswith("s") and k[1:].isdigit() for k in sub_keys):
                fmt_val = "".join([str(value[k]) for k in sub_keys])
            else:
                fmt_val = f"{{{len(value)} EMV tags}}"
                
            # Meaning mapping for compound elements
            if field_key == "de3":
                if fmt_val == "000000":
                    meaning = "Goods and Services (Purchase)"
            elif field_key == "de12":
                s1 = value.get("s1", "")
                s2 = value.get("s2", "")
                if len(s1) == 6 and len(s2) == 6:
                    meaning = f"20{s1[0:2]}-{s1[2:4]}-{s1[4:6]} {s2[0:2]}:{s2[2:4]}:{s2[4:6]}"
            elif field_key == "de43":
                name, city, country = self.get_merchant_name_and_city(tx)
                fmt_val = value.get("s2", "").strip()
                meaning = f"{name} ({city + ', ' if city else ''}{country})"
        else:
            # Simple field
            value_str = str(value)
            fmt_val = value_str
            
            if field_key == "de2":
                fmt_val = self.mask_pan(value_str) if self.is_pan_masked else self.format_pan_spacing(value_str)
                meaning = self.get_card_brand(value_str)
            elif field_key == "de4":
                curr_code = tx.get("de49", "")
                meaning = self.format_amount(value_str, curr_code)
            elif field_key == "de14":
                if len(value_str) == 4:
                    meaning = f"{value_str[2:4]}/20{value_str[0:2]}"
            elif field_key == "de24":
                meaning = self.get_function_code_desc(value_str)
            elif field_key == "de26":
                mcc_info = MCC_CODES.get(value_str)
                if mcc_info:
                    meaning = f"{mcc_info.get('category')} - {mcc_info.get('name')}"
                else:
                    meaning = "Unknown MCC"
            elif field_key == "de49":
                curr_info = CURRENCY_CODES.get(value_str)
                if curr_info:
                    meaning = f"{curr_info.get('name')} ({curr_info.get('code')})"
                else:
                    meaning = "Unknown Currency"
                    
        return label, fmt_val, meaning

    def get_merchant_name_and_city(self, tx):
        de43 = tx.get("de43", {})
        s2 = de43.get("s2", "").strip() if de43.get("s2") else ""
        s3 = de43.get("s3", "").strip() if de43.get("s3") else ""
        country = de43.get("s6", "").strip() if de43.get("s6") else "IQ"
        
        if not s2:
            return "Unknown Merchant", "", country
            
        if s3:
            idx = s2.lower().find(s3.lower())
            if idx != -1:
                name = s2[:idx].strip()
                return name or s2, s3, country
                
        return s2, "", country

    def get_card_brand(self, pan):
        if not pan:
            return "OTHER"
        if pan.startswith("4"):
            return "Visa"
        if pan.startswith(("5", "2")):
            return "Mastercard"
        return "OTHER"

    def mask_pan(self, pan):
        if not pan:
            return ""
        length = len(pan)
        if length < 10:
            return pan
        masked = pan[:6] + "*" * (length - 10) + pan[-4:]
        return self.format_pan_spacing(masked)

    def format_pan_spacing(self, pan):
        if not pan:
            return ""
        # Chunk spacing
        return " ".join([pan[i:i+4] for i in range(0, len(pan), 4)])

    def format_amount(self, amount_str, currency_code):
        try:
            raw_val = float(amount_str)
            meta = CURRENCY_CODES.get(currency_code, {"decimals": 2, "code": currency_code})
            decimals = meta.get("decimals", 2)
            code = meta.get("code", currency_code)
            real_val = raw_val / (10 ** decimals)
            return f"{real_val:,.{decimals}f} {code}"
        except Exception:
            return amount_str

    def get_mti_short_name(self, mti):
        mti_map = {
            "1240": "Presentment",
            "1644": "Fee/Admin",
            "1442": "Chargeback",
            "1120": "Auth Advice"
        }
        return mti_map.get(mti, "Other")

    def get_function_code_desc(self, de24):
        fc_map = {
            "200": "First Presentment",
            "205": "Second Presentment",
            "282": "Chargeback",
            "450": "Second Chargeback",
            "697": "Fee Collection"
        }
        return fc_map.get(de24, f"Func {de24}")

    def hex_to_ascii(self, hex_str):
        try:
            chars = []
            for i in range(0, len(hex_str), 2):
                val = int(hex_str[i:i+2], 16)
                if 32 <= val <= 126:
                    chars.append(chr(val))
            return "".join(chars).strip()
        except Exception:
            return hex_str

    def parse_bcd_date(self, bcd_str):
        if len(bcd_str) == 6:
            return f"20{bcd_str[0:2]}-{bcd_str[2:4]}-{bcd_str[4:6]}"
        elif len(bcd_str) == 4:
            return f"20{bcd_str[0:2]}-{bcd_str[2:4]}"
        return bcd_str

    def parse_emv_amount(self, amount_str):
        try:
            val = int(amount_str)
            return f"{val / 100:,.2f}"
        except Exception:
            return amount_str

if __name__ == "__main__":
    app = IPMViewerApp()
    app.mainloop()
